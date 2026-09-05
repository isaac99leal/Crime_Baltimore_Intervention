#!/usr/bin/env python3
"""Normalize authoritative wine-reference snapshots for the game.

This script does not invent appellations, regions, countries, or grape names.
It converts upstream reference files into compact, source-labelled indices.
"""

from __future__ import annotations

import argparse
import json
import pathlib
import re
import unicodedata
from collections import defaultdict
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import load_workbook

COUNTRY_ALIASES = {
    "Korea, Rep.": "South Korea",
    "Turkiye": "Turkey",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def country_name(value: Any) -> str:
    name = text(value)
    return COUNTRY_ALIASES.get(name, name)


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return None


def dump(path: pathlib.Path, payload: Any, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if compact:
        content = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    else:
        content = json.dumps(payload, ensure_ascii=False, indent=2)
    path.write_text(content + "\n", encoding="utf-8")


def workbook_rows(path: pathlib.Path, sheet: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    for row in rows:
        yield dict(zip(headers, row))


def build_adelaide(snapshot: pathlib.Path, output: pathlib.Path, retrieved_at: str) -> dict[str, int]:
    varieties_path = snapshot / "adelaide_varieties.xlsx"
    origin_colour_path = snapshot / "adelaide_origin_colour.xlsx"
    national_path = snapshot / "adelaide_national.xlsx"
    regional_path = snapshot / "adelaide_regional_2023.xlsx"

    origin_by_prime: dict[str, str] = {}
    for row in workbook_rows(origin_colour_path, "Origin by P"):
        prime = text(row.get("prime"))
        origin = country_name(row.get("porigin"))
        if prime and origin:
            origin_by_prime.setdefault(prime, origin)

    colour_by_prime: dict[str, str] = {}
    for row in workbook_rows(origin_colour_path, "Colour by C and P"):
        prime = text(row.get("prime"))
        colour = text(row.get("pcolour"))
        if prime and colour:
            colour_by_prime.setdefault(prime, colour)

    varieties = []
    for row in workbook_rows(varieties_path, "World alphabetical"):
        prime = text(row.get("prime"))
        if not prime or prime.lower() == "other":
            continue
        varieties.append({
            "name": prime,
            "origin": origin_by_prime.get(prime),
            "colour": colour_by_prime.get(prime),
            "area2000": number(row.get("area2000")),
            "area2010": number(row.get("area2010")),
            "area2016": number(row.get("area2016")),
            "area2023": number(row.get("area2023")),
        })

    countries = []
    for row in workbook_rows(national_path, "All countries - no P"):
        name = country_name(row.get("country"))
        if not name:
            continue
        countries.append({
            "name": name,
            "area2000": number(row.get("area2000")),
            "area2010": number(row.get("area2010")),
            "area2016": number(row.get("area2016")),
            "area2023": number(row.get("area2023")),
        })

    grouped: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    years: dict[tuple[str, str, str, str], set[int]] = defaultdict(set)
    for row in workbook_rows(regional_path, "All countries"):
        country = country_name(row.get("country"))
        prime = text(row.get("prime"))
        if not country or not prime or prime.lower() == "other":
            continue
        region = text(row.get("region"))
        sub_region = text(row.get("sub_region"))
        sub_sub_region = text(row.get("sub_sub_region"))
        key = (country, region, sub_region, sub_sub_region)
        year = row.get("year")
        if isinstance(year, (int, float)):
            years[key].add(int(year))
        grouped[key].append({
            "name": prime,
            "area": number(row.get("area")) or 0,
            "origin": country_name(row.get("porigin")) or None,
            "colour": text(row.get("pcolour")) or None,
            "sourceCode": text(row.get("psource")) or None,
        })

    places = []
    for (country, region, sub_region, sub_sub_region), plantings in grouped.items():
        path = [part for part in (region, sub_region, sub_sub_region) if part]
        top = sorted(plantings, key=lambda item: item["area"], reverse=True)[:8]
        places.append({
            "country": country,
            "path": path,
            "scope": "regional" if path else "national",
            "sourceYears": sorted(years[(country, region, sub_region, sub_sub_region)]),
            "topPlantings": top,
        })

    source = {
        "publisher": "Wine Economics Research Centre, University of Adelaide",
        "title": "Database of Regional, National and Global Winegrape Bearing Areas by Variety, 2000 to 2023",
        "authors": ["Kym Anderson", "Signe Nelgen", "Germán Puga"],
        "edition": "December 2025",
        "url": "https://economics.adelaide.edu.au/wine-economics/databases",
        "retrievedAt": retrieved_at,
        "note": "Statistical planting geography is not automatically an appellation or legal GI.",
    }

    dump(output / "adelaide-varieties.json", {"source": source, "count": len(varieties), "records": varieties}, compact=True)
    dump(output / "adelaide-countries.json", {"source": source, "count": len(countries), "records": countries})
    dump(output / "adelaide-regions.json", {"source": source, "count": len(places), "records": places}, compact=True)
    return {"varieties": len(varieties), "countries": len(countries), "statisticalPlaces": len(places)}


def cfr_metadata(raw: str) -> dict[str, str | None] | None:
    """Parse the CFR cell without losing TTB effective-date annotations."""
    cleaned = unicodedata.normalize("NFKC", raw).replace("\u200b", " ").strip()
    match = re.search(r"\b(9\.\d+)\b", cleaned)
    if not match:
        return None
    effective = re.search(r"effective\s+([0-9/]+)", cleaned, flags=re.IGNORECASE)
    return {
        "cfr": match.group(1),
        "effectiveDate": effective.group(1) if effective else None,
    }


def build_ttb(snapshot: pathlib.Path, output: pathlib.Path, retrieved_at: str) -> int:
    html = (snapshot / "ttb-established-avas.html").read_text(encoding="utf-8")
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    records: list[dict[str, Any]] = []

    if len(tables) < 4:
        raise RuntimeError("TTB AVA page structure changed: expected at least four tables")

    current_state = ""
    for row in tables[1].find_all("tr"):
        cells = [" ".join(cell.stripped_strings).strip() for cell in row.find_all(["th", "td"])]
        if not cells:
            continue
        if len(cells) == 1 and re.fullmatch(r"[A-Z ]+", cells[0]) and cells[0] not in {"BACK TO TOP"}:
            current_state = cells[0].title()
            continue
        metadata = cfr_metadata(cells[-1]) if len(cells) >= 5 else None
        if metadata:
            records.append({
                "name": cells[0],
                "states": [current_state],
                "counties": cells[1],
                "locatedWithin": cells[2],
                "contains": cells[3],
                **metadata,
            })

    for row in tables[3].find_all("tr"):
        cells = [" ".join(cell.stripped_strings).strip() for cell in row.find_all(["th", "td"])]
        metadata = cfr_metadata(cells[-1]) if len(cells) >= 5 else None
        if metadata:
            records.append({
                "name": cells[0],
                "states": [cells[1]],
                "counties": "",
                "locatedWithin": cells[2],
                "contains": cells[3],
                **metadata,
            })

    unique = {record["name"]: record for record in records}
    records = sorted(unique.values(), key=lambda item: item["name"])
    if len(records) != 280:
        raise RuntimeError(f"TTB parser returned {len(records)} AVAs; expected official count 280")

    source = {
        "publisher": "Alcohol and Tobacco Tax and Trade Bureau (TTB)",
        "title": "Established American Viticultural Areas",
        "url": "https://www.ttb.gov/regulated-commodities/beverage-alcohol/wine/established-avas",
        "retrievedAt": retrieved_at,
        "note": "Rows with a future TTB effective date remain marked with effectiveDate and must not be treated as effective before that date.",
    }
    dump(output / "ttb-avas.json", {"source": source, "count": len(records), "records": records}, compact=True)
    return len(records)


def build_eambrosia(snapshot: pathlib.Path, output: pathlib.Path, retrieved_at: str) -> int:
    raw = json.loads((snapshot / "eambrosia-wine-gis.json").read_text(encoding="utf-8"))
    records = []
    for item in raw.get("records", []):
        names = [text(name) for name in item.get("protectedNames", []) if text(name)]
        if not names:
            continue
        records.append({
            "id": item.get("giIdentifier"),
            "names": names,
            "fileNumber": item.get("fileNumber"),
            "countryCodes": item.get("countries") or [],
            "type": item.get("giType"),
            "status": item.get("status"),
            "protectionDate": item.get("euProtectionDate"),
            "modified": item.get("modificationDate"),
        })

    source = {
        "publisher": "European Commission",
        "title": "eAmbrosia Geographical Indications Register — wine records",
        "url": "https://webgate.ec.europa.eu/eambrosia-api/api/v1/geographical-indications",
        "retrievedAt": retrieved_at,
        "note": "Registry membership does not by itself encode the full product specification or authorized grape rules.",
    }
    dump(output / "eambrosia-wine-gis.json", {"source": source, "count": len(records), "records": records}, compact=True)
    return len(records)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("snapshot_dir", type=pathlib.Path)
    parser.add_argument("output_dir", type=pathlib.Path)
    args = parser.parse_args()

    summary = json.loads((args.snapshot_dir / "summary.json").read_text(encoding="utf-8"))
    retrieved_at = summary.get("generatedAt", "")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    counts = build_adelaide(args.snapshot_dir, args.output_dir, retrieved_at)
    counts["ttbAvas"] = build_ttb(args.snapshot_dir, args.output_dir, retrieved_at)
    counts["eambrosiaWineGis"] = build_eambrosia(args.snapshot_dir, args.output_dir, retrieved_at)

    manifest = {
        "generatedAt": retrieved_at,
        "policy": "Reference identity and geography are factual source data. Statistical regions are not relabelled as legal appellations. Generated producers/cuvees remain fictional.",
        "counts": counts,
        "sourceSnapshotErrors": summary.get("errors", []),
    }
    dump(args.output_dir / "manifest.json", manifest)
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
